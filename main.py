from tkinter import *
from tkinter import ttk
from openpyxl import Workbook,load_workbook
from tkcalendar import DateEntry

root = Tk()
root.geometry("600x600")
root.title("Garuda Travels")

def calculate_charges():
    source = sourcevalue.get()
    destination = destinationvalue.get()
    distance = calculate_distance(source, destination)
    charges = distance * rate_per_km
    charges_label.config(text=f"Rs. {charges}")

def calculate_distance(source, destination):
    if source in city_distances and destination in city_distances[source]:
        return city_distances[source][destination]
    elif destination in city_distances and source in city_distances[destination]:
        return city_distances[destination][source]
    else:
        return "N/A"

def submit_details():
    # Get data from the entry fields
    name = namevalue.get()
    phone = phonevalue.get()
    gender = gendervalue.get()
    emergency = emergencyvalue.get()
    payment_mode = paymentmodevalue.get()
    meals = [meal_options[i] for i, var in enumerate(meal_values) if var.get()]
    source = sourcevalue.get()
    destination = destinationvalue.get()
    travel_date = cal.get()

    # Calculate charges
    distance = calculate_distance(source, destination)
    charges = distance * rate_per_km

    # Create or load workbook
    try:
        wb = load_workbook(r"C:\Users\moham\PycharmProjects\TravelApplicationForm\TravellingAgency.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    # Append data to the worksheet
    ws.append([name, phone, gender, emergency, payment_mode, ", ".join(meals), source, destination, travel_date, charges])

    # Save the workbook
    wb.save(r"C:\Users\moham\PycharmProjects\TravelApplicationForm\TravellingAgency.xlsx")
    print("Data saved successfully!")



mainframe = ttk.Frame(root, padding="20")
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
mainframe.columnconfigure(0, weight=1)
mainframe.rowconfigure(0, weight=1)

ttk.Label(mainframe, text="Travel Application Form", font=("Helvetica", 18, "bold")).grid(column=1, row=1, columnspan=2, pady=(0, 10))

ttk.Label(mainframe, text="Name").grid(row=2, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Phone Number").grid(row=3, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Gender").grid(row=4, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Emergency Call").grid(row=5, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Payment Mode").grid(row=6, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Meals").grid(row=7, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Source").grid(row=8, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Destination").grid(row=9, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Travel Date").grid(row=10, column=1, sticky=E, pady=5)
ttk.Label(mainframe, text="Total Booking Charges:").grid(row=11, column=1, sticky=E, pady=5)

namevalue = StringVar()
phonevalue = StringVar()
gendervalue = StringVar()
emergencyvalue = StringVar()
paymentmodevalue = StringVar()
meal_values = [BooleanVar() for _ in range(7)]
foodservicevalue = BooleanVar()
sourcevalue = StringVar()
destinationvalue = StringVar()

name_entry = ttk.Entry(mainframe, textvariable=namevalue, width=30)
name_entry.grid(row=2, column=2, sticky=W)

phone_entry = ttk.Entry(mainframe, textvariable=phonevalue, width=30)
phone_entry.grid(row=3, column=2, sticky=W)

gender_combobox = ttk.Combobox(mainframe, textvariable=gendervalue, values=["Male", "Female", "Other"], width=28)
gender_combobox.grid(row=4, column=2, sticky=W)

emergency_entry = ttk.Entry(mainframe, textvariable=emergencyvalue, width=30)
emergency_entry.grid(row=5, column=2, sticky=W)

paymentmode_combobox = ttk.Combobox(mainframe, textvariable=paymentmodevalue, values=["On Cash", "UPI", "Credit/Debit card payment"], width=28)
paymentmode_combobox.grid(row=6, column=2, sticky=W)

meal_options = ["Italian", "Indian", "Chinese", "Mexican", "Korean", "Japanese", "American"]
meal_frame = ttk.Frame(mainframe)
meal_frame.grid(row=7, column=2, sticky=W)
for i, meal in enumerate(meal_options):
    ttk.Checkbutton(meal_frame, text=meal, variable=meal_values[i]).grid(row=0, column=i, sticky=W)

source_options = ["New York", "Los Angeles", "Chicago", "San Francisco", "Miami", "Delhi", "Mumbai", "Hyderabad", "Kolkata", "Bangalore", "Chennai", "Mysore", "Kurnool", "Ananthapur", "Guntur", "Adilabad", "Mahboob Nagar", "Adoni", "Raichur", "Pune"]
source_combobox = ttk.Combobox(mainframe, textvariable=sourcevalue, values=source_options, width=28)
source_combobox.grid(row=8, column=2, sticky=W)

destination_options = ["London", "Paris", "Tokyo", "Dubai", "Sydney", "Delhi", "Mumbai", "Hyderabad", "Kolkata", "Bangalore", "Chennai", "Mysore", "Kurnool", "Ananthapur", "Guntur", "Adilabad", "Mahboob Nagar", "Adoni", "Raichur", "Pune"]
destination_combobox = ttk.Combobox(mainframe, textvariable=destinationvalue, values=destination_options, width=28)
destination_combobox.grid(row=9, column=2, sticky=W)

cal = DateEntry(mainframe, textvariable=StringVar(), width=28, background='darkblue', foreground='white', borderwidth=2, year=2022)
cal.grid(row=10, column=2, sticky=W)

charges_label = ttk.Label(mainframe, text="")
charges_label.grid(row=11, column=2, sticky=W)

calculate_button = ttk.Button(mainframe, text="Calculate Charges", command=calculate_charges)
calculate_button.grid(row=12, columnspan=2, pady=10)

submit_button = ttk.Button(mainframe, text="Submit", command=submit_details)
submit_button.grid(row=13, columnspan=2)

for child in mainframe.winfo_children():
    child.grid_configure(padx=10, pady=5)

# Dictionary mapping cities to their distances from each other (in km)
city_distances = {
    "New York": {"London": 5576, "Paris": 5839, "Tokyo": 10841, "Dubai": 11001, "Sydney": 16014, "Delhi": 11755, "Mumbai": 12047, "Hyderabad": 1194, "Kolkata": 1420, "Bangalore": 3414, "Chennai": 1143, "Mysore": 3370, "Kurnool": 1767, "Ananthapur": 1461, "Guntur": 1106, "Adilabad": 579, "Mahboob Nagar": 618, "Adoni": 1063, "Raichur": 1161, "Pune": 1202},
    "Los Angeles": {"London": 8797, "Paris": 9079, "Tokyo": 8811, "Dubai": 13416, "Sydney": 12051, "Delhi": 11967, "Mumbai": 12041, "Hyderabad": 13314, "Kolkata": 12694, "Bangalore": 14173, "Chennai": 14169, "Mysore": 13797, "Kurnool": 14367, "Ananthapur": 14749, "Guntur": 13470, "Adilabad": 11525, "Mahboob Nagar": 11293, "Adoni": 14085, "Raichur": 14832, "Pune": 13986},
    "Chicago": {"London": 6165, "Paris": 6415, "Tokyo": 10922, "Dubai": 10856, "Sydney": 15694, "Delhi": 11418, "Mumbai": 11436, "Hyderabad": 12779, "Kolkata": 12966, "Bangalore": 13854, "Chennai": 13874, "Mysore": 13497, "Kurnool": 14094, "Ananthapur": 14443, "Guntur": 13188, "Adilabad": 11432, "Mahboob Nagar": 11201, "Adoni": 13834, "Raichur": 14581, "Pune": 13735},
    "San Francisco": {"London": 8618, "Paris": 8885, "Tokyo": 8971, "Dubai": 13446, "Sydney": 12068, "Delhi": 12346, "Mumbai": 12467, "Hyderabad": 12842, "Kolkata": 13323, "Bangalore": 13520, "Chennai": 14166, "Mysore": 13584, "Kurnool": 14145, "Ananthapur": 14508, "Guntur": 13214, "Adilabad": 11220, "Mahboob Nagar": 10988, "Adoni": 13680, "Raichur": 14427, "Pune": 13481},
    "Miami": {"London": 7054, "Paris": 7296, "Tokyo": 14769, "Dubai": 12249, "Sydney": 16911, "Delhi": 11913, "Mumbai": 11744, "Hyderabad": 13268, "Kolkata": 13720, "Bangalore": 14536, "Chennai": 15208, "Mysore": 14742, "Kurnool": 15313, "Ananthapur": 15676, "Guntur": 14382, "Adilabad": 12387, "Mahboob Nagar": 12155, "Adoni": 14847, "Raichur": 15594, "Pune": 14648},
    "Delhi": {"Mumbai": 1148, "Hyderabad": 1253, "Kolkata": 1486, "Bangalore": 1750, "Chennai": 1765, "Mysore": 1846, "Kurnool": 984, "Ananthapur": 903, "Guntur": 1293, "Adilabad": 1287, "Mahboob Nagar": 1302, "Adoni": 896, "Raichur": 827, "Pune": 1175},
    "Mumbai": {"Hyderabad": 621, "Kolkata": 1707, "Bangalore": 841, "Chennai": 1248, "Mysore": 984, "Kurnool": 1507, "Ananthapur": 1586, "Guntur": 658, "Adilabad": 668, "Mahboob Nagar": 784, "Adoni": 1194, "Raichur": 1230, "Pune": 172},
    "Hyderabad": {"Kolkata": 1464, "Bangalore": 569, "Chennai": 619, "Mysore": 830, "Kurnool": 219, "Ananthapur": 342, "Guntur": 291, "Adilabad": 305, "Mahboob Nagar": 90, "Adoni": 246, "Raichur": 282, "Pune": 544},
    "Kolkata": {"Bangalore": 1866, "Chennai": 1366, "Mysore": 1791, "Kurnool": 1678, "Ananthapur": 1559, "Guntur": 1094, "Adilabad": 1251, "Mahboob Nagar": 1210, "Adoni": 1399, "Raichur": 1418, "Pune": 2023},
    "Bangalore": {"Chennai": 347, "Mysore": 139, "Kurnool": 409, "Ananthapur": 334, "Guntur": 548, "Adilabad": 664, "Mahboob Nagar": 768, "Adoni": 574, "Raichur": 641, "Pune": 736},
    "Chennai": {"Mysore": 500, "Kurnool": 678, "Ananthapur": 719, "Guntur": 426, "Adilabad": 646, "Mahboob Nagar": 752, "Adoni": 689, "Raichur": 724, "Pune": 1213},
    "Mysore": {"Kurnool": 643, "Ananthapur": 759, "Guntur": 608, "Adilabad": 795, "Mahboob Nagar": 901, "Adoni": 713, "Raichur": 743, "Pune": 900},
    "Kurnool": {"Ananthapur": 129, "Guntur": 393, "Adilabad": 502, "Mahboob Nagar": 491, "Adoni": 160, "Raichur": 102, "Pune": 632},
    "Ananthapur": {"Guntur": 267, "Adilabad": 557, "Mahboob Nagar": 445, "Adoni": 269, "Raichur": 209, "Pune": 744},
    "Guntur": {"Adilabad": 521, "Mahboob Nagar": 619, "Adoni": 337, "Raichur": 272, "Pune": 816},
    "Adilabad": {"Mahboob Nagar": 165, "Adoni": 319, "Raichur": 355, "Pune": 739},
    "Mahboob Nagar": {"Adoni": 179, "Raichur": 120, "Pune": 664},
    "Adoni": {"Raichur": 42, "Pune": 598},
    "Raichur": {"Pune": 556}
}

# Rate per kilometer (in Rs.)
rate_per_km = 25

root.mainloop()
